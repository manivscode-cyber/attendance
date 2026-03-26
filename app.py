from __future__ import annotations

import base64
import hmac
import os
import pickle
import re
import secrets
import sqlite3
import smtplib
import sys
from collections import defaultdict, deque
from datetime import date, datetime, time, timedelta
from email.message import EmailMessage
from functools import wraps
from io import BytesIO
from pathlib import Path
from threading import Lock
from time import monotonic
from typing import Any
from urllib.parse import urlencode, urlsplit
from urllib.request import Request, urlopen

import numpy as np
import pytz
from PIL import Image
from flask_wtf.csrf import CSRFError, CSRFProtect, generate_csrf
from flask import (
    Flask,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash

BASE_DIR = Path(__file__).resolve().parent
DATABASE_PATH = BASE_DIR / "attendance.db"
STATIC_DIR = BASE_DIR / "static"
UPLOAD_DIR = STATIC_DIR / "uploads" / "employees"
VENV_PYTHON = BASE_DIR / ".venv" / "Scripts" / "python.exe"
IST = pytz.timezone("Asia/Kolkata")
CHECKIN_CUTOFF = time(10, 0)
DEPARTMENT_OPTIONS = ["IT", "Editor", "Marketing", "Sales", "Non IT"]
WORK_MODE_OPTIONS = ["WFH", "Onsite", "Field"]
VALID_STATUS_OPTIONS = {"On Time", "Late", "Pending"}
VALID_ATTENDANCE_TYPE_OPTIONS = {
    "In Progress",
    "Half Day",
    "Full Day",
    "Pending",
}
NAME_PATTERN = re.compile(r"[A-Za-z][A-Za-z .'\-]{1,79}")
EMAIL_PATTERN = re.compile(
    r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$"
)
MANAGER_PASSWORD_SETTING = "manager_password_hash"
MANAGER_EMAIL_SETTING = "manager_recovery_email"
MANAGER_PHONE_SETTING = "manager_recovery_phone"
MANAGER_RESET_CODE_HASH_SETTING = "manager_reset_code_hash"
MANAGER_RESET_CODE_EXPIRY_SETTING = "manager_reset_code_expires_at"
MANAGER_RESET_CODE_EMAIL_SETTING = "manager_reset_code_email"
MANAGER_RESET_CODE_TARGET_SETTING = "manager_reset_code_target"
MANAGER_RESET_CODE_CHANNEL_SETTING = "manager_reset_code_channel"
MANAGER_SESSION_KEY = "_manager_authenticated_at"
MANAGER_RESET_VERIFIED_SESSION_KEY = "_manager_reset_verified"
MANAGER_RESET_RESEND_SESSION_KEY = "_manager_reset_resend_after"
MANAGER_RESET_SUCCESS_SESSION_KEY = "_manager_reset_success"
PENDING_SCAN_SESSION_KEY = "_pending_scan_employee"
MANAGER_SESSION_DURATION = timedelta(hours=6)
MANAGER_RESET_VERIFIED_TTL = timedelta(minutes=15)
MANAGER_RESET_RESEND_COOLDOWN = timedelta(seconds=30)
PENDING_SCAN_TTL = timedelta(minutes=5)
MANAGER_RESET_CODE_TTL = timedelta(minutes=10)
MANAGER_RESET_CODE_LENGTH = 6
MAX_FACE_IMAGE_BYTES = 6 * 1024 * 1024
MAX_IMAGE_DATA_LENGTH = MAX_FACE_IMAGE_BYTES * 2
ALLOWED_IMAGE_MIME_TYPES = {"image/jpeg", "image/png", "image/webp"}
MANAGER_PASSWORD_ONLY_MESSAGE = (
    "Forgot password is disabled. Use the manager password to sign in."
)
MANAGER_RECOVERY_DISABLED_MESSAGE = (
    "Recovery settings are disabled because forgot password was removed."
)
RATE_LIMIT_BUCKETS: dict[str, deque[float]] = defaultdict(deque)
RATE_LIMIT_LOCK = Lock()
Image.MAX_IMAGE_PIXELS = 20_000_000


def load_local_env_file(path: Path) -> None:
    if not path.exists():
        return

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        env_key = key.strip()
        if not env_key or env_key in os.environ:
            continue

        env_value = value.strip()
        if len(env_value) >= 2 and env_value[0] == env_value[-1]:
            if env_value[0] in {'"', "'"}:
                env_value = env_value[1:-1]
        os.environ[env_key] = env_value


load_local_env_file(BASE_DIR / ".env")


def add_local_venv_site_packages() -> None:
    candidate_paths = [
        BASE_DIR / ".venv" / "Lib" / "site-packages",
    ]
    candidate_paths.extend(
        (BASE_DIR / ".venv" / "lib").glob("python*/site-packages")
        if (BASE_DIR / ".venv" / "lib").exists()
        else []
    )

    for path in candidate_paths:
        if path.exists():
            resolved = str(path.resolve())
            if resolved not in sys.path:
                sys.path.insert(0, resolved)


def face_stack_runtime_hint() -> str:
    current_python = f"{sys.executable} (Python {sys.version.split()[0]})"
    if VENV_PYTHON.exists():
        return (
            "Start the app with the project virtual environment: "
            f"{VENV_PYTHON}. Current interpreter: {current_python}"
        )
    return f"Current interpreter: {current_python}"


try:
    import cv2
except Exception:  # pragma: no cover - handled at runtime
    add_local_venv_site_packages()
    try:
        import cv2
    except Exception as retry_exc:  # pragma: no cover - handled at runtime
        cv2 = None
        CV2_IMPORT_ERROR = str(retry_exc)
    else:
        CV2_IMPORT_ERROR = None
else:
    CV2_IMPORT_ERROR = None

try:
    import face_recognition
except Exception:  # pragma: no cover - handled at runtime
    add_local_venv_site_packages()
    try:
        import face_recognition
    except Exception as retry_exc:  # pragma: no cover - handled at runtime
        face_recognition = None
        FACE_IMPORT_ERROR = str(retry_exc)
    else:
        FACE_IMPORT_ERROR = None
else:
    FACE_IMPORT_ERROR = None

app = Flask(__name__, template_folder="templates", static_folder="static")
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
app.config.update(
    SECRET_KEY=os.getenv("GODIGITAL_SECRET_KEY") or secrets.token_urlsafe(32),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_NAME="godigital_attendance_session",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=10),
    MAX_CONTENT_LENGTH=8 * 1024 * 1024,
    WTF_CSRF_FIELD_NAME="_csrf_token",
    WTF_CSRF_HEADERS=["X-CSRFToken", "X-CSRF-Token"],
    WTF_CSRF_TIME_LIMIT=None,
)
csrf = CSRFProtect(app)


def now_ist() -> datetime:
    return datetime.now(IST)


def today_ist() -> date:
    return now_ist().date()


def iso_now() -> str:
    return now_ist().isoformat(timespec="seconds")


def current_system_label() -> str:
    return now_ist().strftime("%d %b %Y, %I:%M %p IST")


def current_date_label(day: date | None = None) -> str:
    return (day or today_ist()).strftime("%B %d, %Y")


def short_date_label(day: date | None = None) -> str:
    return (day or today_ist()).strftime("%B %d")


def parse_business_date(
    value: str | None,
    fallback: date | None = None,
) -> date:
    if value:
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            pass
    return fallback or today_ist()


def localize_ist(dt: datetime) -> datetime:
    if dt.tzinfo is not None:
        return dt.astimezone(IST)
    return IST.localize(dt)


def parse_stored_datetime(
    value: str | None,
    business_date: date | None = None,
) -> datetime | None:
    if not value:
        return None

    text = str(value).strip()
    if not text:
        return None

    normalized = text.replace("Z", "+00:00")
    try:
        return localize_ist(datetime.fromisoformat(normalized))
    except ValueError:
        pass

    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
    ):
        try:
            return localize_ist(datetime.strptime(text, fmt))
        except ValueError:
            continue

    target_date = business_date or today_ist()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            parsed_time = datetime.strptime(text, fmt).time()
            return localize_ist(datetime.combine(target_date, parsed_time))
        except ValueError:
            continue

    return None


def build_iso_timestamp(
    business_date: date,
    clock_value: str | None,
) -> str | None:
    if not clock_value:
        return None

    text = clock_value.strip()
    if not text:
        return None

    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            parsed_time = datetime.strptime(text, fmt).time()
            return localize_ist(
                datetime.combine(business_date, parsed_time)
            ).isoformat(timespec="seconds")
        except ValueError:
            continue

    parsed = parse_stored_datetime(text, business_date)
    if parsed:
        return parsed.isoformat(timespec="seconds")
    return None


def compute_total_hours(
    check_in_value: str | None,
    check_out_value: str | None,
    business_date: date,
) -> float:
    check_in_dt = parse_stored_datetime(check_in_value, business_date)
    check_out_dt = parse_stored_datetime(check_out_value, business_date)
    if not check_in_dt or not check_out_dt or check_out_dt <= check_in_dt:
        return 0.0

    return round((check_out_dt - check_in_dt).total_seconds() / 3600, 2)


def compute_live_hours(
    check_in_value: str | None,
    check_out_value: str | None,
    business_date: date,
) -> float:
    check_in_dt = parse_stored_datetime(check_in_value, business_date)
    if not check_in_dt:
        return 0.0

    check_out_dt = parse_stored_datetime(check_out_value, business_date)
    if not check_out_dt:
        check_out_dt = now_ist() if business_date == today_ist() else None

    if not check_out_dt or check_out_dt <= check_in_dt:
        return 0.0

    return round((check_out_dt - check_in_dt).total_seconds() / 3600, 2)


def hours_to_label(total_hours: float | None) -> str:
    if not total_hours:
        return "0h 0m"

    hours = int(total_hours)
    minutes = int(round((total_hours - hours) * 60))
    if minutes == 60:
        hours += 1
        minutes = 0
    return f"{hours}h {minutes}m"


def normalize_department(value: str | None) -> str:
    cleaned = " ".join((value or "").split()).strip()
    if not cleaned:
        return DEPARTMENT_OPTIONS[0]

    mapping = {option.casefold(): option for option in DEPARTMENT_OPTIONS}
    return mapping.get(cleaned.casefold(), cleaned)


def normalize_work_mode(value: str | None) -> str:
    cleaned = " ".join((value or "").split()).strip()
    if not cleaned:
        return "Onsite"

    mapping = {option.casefold(): option for option in WORK_MODE_OPTIONS}
    return mapping.get(cleaned.casefold(), cleaned)


def normalize_optional_work_mode(value: str | None) -> str | None:
    cleaned = " ".join((value or "").split()).strip()
    if not cleaned:
        return None
    return normalize_work_mode(cleaned)


def normalize_person_name(value: str | None) -> str:
    return " ".join((value or "").split()).strip().casefold()


def normalize_phone_lookup(value: str | None) -> str:
    return re.sub(r"\D+", "", value or "")


def clean_employee_name(value: str | None) -> str:
    return " ".join((value or "").split()).strip()


def validate_employee_name(value: str | None) -> tuple[str | None, str | None]:
    cleaned = clean_employee_name(value)
    if not cleaned:
        return None, "Employee name is required."
    if not NAME_PATTERN.fullmatch(cleaned):
        return (
            None,
            "Employee name should be 2 to 80 characters "
            "and contain only letters, spaces, dots, "
            "apostrophes, or hyphens.",
        )
    return cleaned, None


def validate_phone_number(
    value: str | None,
    *,
    required: bool = True,
) -> tuple[str | None, str | None]:
    normalized = normalize_phone_lookup(value)
    if not normalized:
        if required:
            return None, "Phone number must contain exactly 10 digits."
        return "", None
    if len(normalized) != 10:
        return None, "Phone number must contain exactly 10 digits."
    return normalized, None


def normalize_email(value: str | None) -> str:
    return " ".join((value or "").split()).strip().lower()


def validate_email_address(value: str | None) -> tuple[str | None, str | None]:
    cleaned = normalize_email(value)
    if not cleaned:
        return None, "Recovery email is required."
    if not EMAIL_PATTERN.fullmatch(cleaned):
        return None, "Enter a valid recovery email address."
    return cleaned, None


def normalize_recovery_phone(value: str | None) -> str:
    cleaned = " ".join((value or "").split()).strip()
    if not cleaned:
        return ""

    if cleaned.startswith("+"):
        digits = normalize_phone_lookup(cleaned[1:])
        return f"+{digits}" if digits else ""
    return normalize_phone_lookup(cleaned)


def validate_recovery_phone(
    value: str | None,
    *,
    required: bool = False,
) -> tuple[str | None, str | None]:
    normalized = normalize_recovery_phone(value)
    if not normalized:
        if required:
            return None, "Enter a valid recovery phone number."
        return "", None

    digits_only = normalize_phone_lookup(normalized)
    if len(digits_only) < 10 or len(digits_only) > 15:
        return None, "Recovery phone should contain 10 to 15 digits."
    return normalized, None


def validate_manager_recovery_inputs(
    email_value: str | None,
    phone_value: str | None,
    *,
    require_any: bool,
) -> tuple[str, str, str | None]:
    email = ""
    phone = ""

    if normalize_email(email_value):
        validated_email, email_error = validate_email_address(email_value)
        if email_error or not validated_email:
            return "", "", email_error
        email = validated_email

    if normalize_recovery_phone(phone_value):
        validated_phone, phone_error = validate_recovery_phone(phone_value)
        if phone_error or validated_phone is None:
            return "", "", phone_error
        phone = validated_phone

    if require_any and not email and not phone:
        return "", "", "Add at least one recovery email or phone number."

    return email, phone, None


def validate_manager_recovery_contact(
    value: str | None,
) -> tuple[str | None, str | None, str | None]:
    cleaned = " ".join((value or "").split()).strip()
    if not cleaned:
        return None, None, "Enter the recovery email or phone number."

    if "@" in cleaned:
        email, email_error = validate_email_address(cleaned)
        return email, "email", email_error

    phone, phone_error = validate_recovery_phone(cleaned, required=True)
    return phone, "phone", phone_error


def normalize_status_input(value: str | None) -> str | None:
    cleaned = " ".join((value or "").split()).strip()
    if not cleaned:
        return None
    return cleaned if cleaned in VALID_STATUS_OPTIONS else None


def normalize_attendance_type_input(value: str | None) -> str | None:
    cleaned = " ".join((value or "").split()).strip()
    if not cleaned:
        return None
    return cleaned if cleaned in VALID_ATTENDANCE_TYPE_OPTIONS else None


def status_from_checkin(check_in_dt: datetime | None) -> str:
    if not check_in_dt:
        return "Pending"
    return (
        "On Time"
        if check_in_dt.astimezone(IST).time() <= CHECKIN_CUTOFF
        else "Late"
    )


def attendance_type_from_hours(
    total_hours: float,
    has_checkout: bool,
) -> str:
    if not has_checkout:
        return "In Progress"
    return "Full Day" if total_hours > 4.0 else "Half Day"


def badge_tone_for_status(status: str | None) -> str:
    normalized = (status or "").strip().lower()
    if normalized == "on time":
        return "success"
    if normalized == "late":
        return "danger"
    if normalized == "pending":
        return "muted"
    return "info"


def badge_tone_for_attendance_type(attendance_type: str | None) -> str:
    normalized = (attendance_type or "").strip().lower()
    if normalized == "full day":
        return "success"
    if normalized == "half day":
        return "warning"
    if normalized == "in progress":
        return "info"
    return "muted"


def badge_tone_for_work_mode(work_mode: str | None) -> str:
    normalized = (work_mode or "").strip().lower()
    if normalized == "onsite":
        return "success"
    if normalized == "wfh":
        return "info"
    if normalized == "field":
        return "warning"
    return "muted"


def format_time_label(value: str | None, business_date: date) -> str:
    dt = parse_stored_datetime(value, business_date)
    return dt.strftime("%I:%M %p") if dt else "--"


def format_datetime_label(
    value: str | None,
    business_date: date | None = None,
) -> str:
    dt = parse_stored_datetime(value, business_date)
    return dt.strftime("%d %b %Y, %I:%M %p IST") if dt else "--"


def time_input_value(value: str | None, business_date: date) -> str:
    dt = parse_stored_datetime(value, business_date)
    return dt.strftime("%H:%M") if dt else ""


def slugify(value: str) -> str:
    return (
        re.sub(r"[^a-z0-9]+", "-", value.strip().lower()).strip("-")
        or "employee"
    )


def ensure_upload_dir() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        connection = sqlite3.connect(DATABASE_PATH)
        connection.row_factory = sqlite3.Row
        g.db = connection
    return g.db


@app.teardown_appcontext
def close_db(_error: BaseException | None) -> None:
    connection = g.pop("db", None)
    if connection is not None:
        connection.close()


def query_all(query: str, params: tuple[Any, ...] = ()) -> list[sqlite3.Row]:
    return get_db().execute(query, params).fetchall()


def query_one(query: str, params: tuple[Any, ...] = ()) -> sqlite3.Row | None:
    return get_db().execute(query, params).fetchone()


def execute_db(query: str, params: tuple[Any, ...] = ()) -> sqlite3.Cursor:
    cursor = get_db().execute(query, params)
    get_db().commit()
    return cursor


def table_columns(table_name: str) -> set[str]:
    rows = query_all(f"PRAGMA table_info({table_name})")
    return {row["name"] for row in rows}


def add_column_if_missing(
    table_name: str,
    column_name: str,
    column_sql: str,
) -> None:
    if column_name in table_columns(table_name):
        return

    get_db().execute(
        f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_sql}"
    )
    get_db().commit()


def ensure_schema() -> None:
    ensure_upload_dir()

    db = get_db()
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            email TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            department TEXT DEFAULT 'General',
            work_mode TEXT DEFAULT 'Onsite',
            password TEXT NOT NULL DEFAULT '',
            role TEXT NOT NULL DEFAULT 'employee',
            face_encoding BLOB,
            photo_path TEXT,
            created_at TEXT NOT NULL
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id TEXT NOT NULL,
            date TEXT NOT NULL,
            check_in TEXT,
            break_in TEXT,
            break_out TEXT,
            check_out TEXT,
            work_mode TEXT DEFAULT '',
            total_hours REAL DEFAULT 0,
            status TEXT DEFAULT 'Pending',
            attendance_type TEXT DEFAULT 'Pending',
            UNIQUE(employee_id, date)
        )
        """
    )
    db.commit()

    add_column_if_missing("employees", "email", "TEXT DEFAULT ''")
    add_column_if_missing("employees", "phone", "TEXT DEFAULT ''")
    add_column_if_missing("employees", "department", "TEXT DEFAULT 'General'")
    add_column_if_missing("employees", "work_mode", "TEXT DEFAULT 'Onsite'")
    add_column_if_missing("employees", "password", "TEXT DEFAULT ''")
    add_column_if_missing("employees", "role", "TEXT DEFAULT 'employee'")
    add_column_if_missing("employees", "face_encoding", "BLOB")
    add_column_if_missing("employees", "photo_path", "TEXT")
    add_column_if_missing("employees", "created_at", "TEXT")

    add_column_if_missing("attendance", "break_in", "TEXT")
    add_column_if_missing("attendance", "break_out", "TEXT")
    add_column_if_missing("attendance", "work_mode", "TEXT DEFAULT ''")
    add_column_if_missing("attendance", "total_hours", "REAL DEFAULT 0")
    add_column_if_missing("attendance", "status", "TEXT DEFAULT 'Pending'")
    add_column_if_missing(
        "attendance",
        "attendance_type",
        "TEXT DEFAULT 'Pending'",
    )

    db.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_employees_name
        ON employees(name)
        """
    )
    db.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_attendance_date
        ON attendance(date)
        """
    )
    try:
        db.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_attendance_employee_date
            ON attendance(employee_id, date)
            """
        )
    except sqlite3.IntegrityError:
        pass

    db.execute(
        """
        UPDATE employees
        SET department = COALESCE(NULLIF(TRIM(department), ''), 'General'),
            work_mode = COALESCE(NULLIF(TRIM(work_mode), ''), 'Onsite'),
            phone = COALESCE(phone, ''),
            email = COALESCE(email, ''),
            password = COALESCE(password, ''),
            role = COALESCE(NULLIF(TRIM(role), ''), 'employee'),
            created_at = COALESCE(created_at, ?)
        """,
        (iso_now(),),
    )
    db.execute(
        """
        UPDATE attendance
        SET work_mode = COALESCE(
            NULLIF(TRIM(work_mode), ''),
            (
                SELECT COALESCE(NULLIF(TRIM(e.work_mode), ''), 'Onsite')
                FROM employees e
                WHERE e.employee_id = attendance.employee_id
            ),
            'Onsite'
        )
        """
    )
    db.commit()

    missing_ids = query_all(
        """
        SELECT id
        FROM employees
        WHERE employee_id IS NULL OR TRIM(employee_id) = ''
        """
    )
    for row in missing_ids:
        db.execute(
            "UPDATE employees SET employee_id = ? WHERE id = ?",
            (f"GDLEGACY{row['id']:03d}", row["id"]),
        )
    db.commit()


def app_setting(key: str) -> str | None:
    row = query_one(
        """
        SELECT value
        FROM app_settings
        WHERE key = ?
        """,
        (key,),
    )
    return str(row["value"]) if row else None


def set_app_setting(key: str, value: str) -> None:
    execute_db(
        """
        INSERT INTO app_settings (key, value, updated_at)
        VALUES (?, ?, ?)
        ON CONFLICT(key) DO UPDATE SET
            value = excluded.value,
            updated_at = excluded.updated_at
        """,
        (key, value, iso_now()),
    )


def delete_app_setting(key: str) -> None:
    execute_db("DELETE FROM app_settings WHERE key = ?", (key,))


def photo_url(photo_path: str | None) -> str | None:
    if not photo_path:
        return None
    return url_for("static", filename=photo_path)


def static_asset_url(filename: str) -> str:
    asset_path = STATIC_DIR / filename
    version = (
        int(asset_path.stat().st_mtime)
        if asset_path.exists()
        else int(now_ist().timestamp())
    )
    return url_for("static", filename=filename, v=version)


def csrf_token() -> str:
    return generate_csrf()


def client_rate_limit_key(scope: str) -> str:
    return f"{scope}:{request.remote_addr or 'unknown'}"


def consume_rate_limit(
    scope: str,
    max_requests: int,
    window_seconds: int,
) -> int | None:
    current = monotonic()
    bucket_key = client_rate_limit_key(scope)
    with RATE_LIMIT_LOCK:
        bucket = RATE_LIMIT_BUCKETS[bucket_key]
        while bucket and current - bucket[0] >= window_seconds:
            bucket.popleft()
        if len(bucket) >= max_requests:
            return max(1, int(window_seconds - (current - bucket[0])))
        bucket.append(current)
    return None


def rate_limited_json_response(
    message: str,
    retry_after: int,
) -> tuple[Any, int, dict[str, str]]:
    return (
        jsonify(
            {
                "success": False,
                "message": message,
                "retry_after": retry_after,
            }
        ),
        429,
        {"Retry-After": str(retry_after)},
    )


def manager_password_configured() -> bool:
    return bool(
        os.getenv("GODIGITAL_MANAGER_PASSWORD_HASH")
        or os.getenv("GODIGITAL_MANAGER_PASSWORD")
        or app_setting(MANAGER_PASSWORD_SETTING)
    )


def manager_password_managed_by_env() -> bool:
    return bool(
        os.getenv("GODIGITAL_MANAGER_PASSWORD_HASH")
        or os.getenv("GODIGITAL_MANAGER_PASSWORD")
    )


def manager_recovery_email() -> str | None:
    configured = (
        os.getenv("GODIGITAL_MANAGER_RECOVERY_EMAIL")
        or app_setting(MANAGER_EMAIL_SETTING)
    )
    email, _error = validate_email_address(configured)
    return email


def manager_recovery_phone() -> str | None:
    configured = (
        os.getenv("GODIGITAL_MANAGER_RECOVERY_PHONE")
        or app_setting(MANAGER_PHONE_SETTING)
    )
    phone, _error = validate_recovery_phone(configured)
    return phone


def manager_has_recovery_contact() -> bool:
    return bool(manager_recovery_email() or manager_recovery_phone())


def manager_recovery_contacts_managed_by_env() -> bool:
    return bool(
        os.getenv("GODIGITAL_MANAGER_RECOVERY_EMAIL")
        or os.getenv("GODIGITAL_MANAGER_RECOVERY_PHONE")
    )


def manager_reset_email_delivery_ready() -> bool:
    return bool(
        os.getenv("GODIGITAL_SMTP_HOST")
        and os.getenv("GODIGITAL_SMTP_FROM_EMAIL")
    )


def manager_reset_phone_delivery_ready() -> bool:
    return bool(
        os.getenv("GODIGITAL_TWILIO_ACCOUNT_SID")
        and os.getenv("GODIGITAL_TWILIO_AUTH_TOKEN")
        and os.getenv("GODIGITAL_TWILIO_FROM_PHONE")
    )


def manager_reset_email_delivery_available() -> bool:
    return manager_reset_email_delivery_ready()


def manager_reset_phone_delivery_available() -> bool:
    return manager_reset_phone_delivery_ready()


def mask_email_address(value: str | None) -> str:
    email = normalize_email(value)
    if "@" not in email:
        return ""
    local_part, domain = email.split("@", 1)
    prefix = local_part[:2] if len(local_part) > 2 else local_part[:1]
    masked_local = f"{prefix}{'*' * max(1, len(local_part) - len(prefix))}"
    return f"{masked_local}@{domain}"


def mask_phone_number(value: str | None) -> str:
    normalized = normalize_recovery_phone(value)
    digits_only = normalize_phone_lookup(normalized)
    if not digits_only:
        return ""

    suffix = digits_only[-2:]
    prefix = "+" if normalized.startswith("+") else ""
    return f"{prefix}{'*' * max(4, len(digits_only) - 2)}{suffix}"


def mask_manager_contact(value: str | None, channel: str | None) -> str:
    if channel == "phone":
        return mask_phone_number(value)
    return mask_email_address(value)


def recovery_phone_matches(
    configured_phone: str | None,
    entered_phone: str | None,
) -> bool:
    return normalize_phone_lookup(configured_phone) == normalize_phone_lookup(
        entered_phone
    )


def phone_for_sms_delivery(value: str) -> str:
    normalized = normalize_recovery_phone(value)
    if normalized.startswith("+"):
        return normalized

    digits_only = normalize_phone_lookup(normalized)
    if len(digits_only) > 10:
        return f"+{digits_only}"

    default_country_code = (
        os.getenv("GODIGITAL_SMS_DEFAULT_COUNTRY_CODE") or "+91"
    ).strip()
    if not default_country_code.startswith("+"):
        default_country_code = (
            f"+{normalize_phone_lookup(default_country_code)}"
        )
    return f"{default_country_code}{digits_only}"


def save_manager_recovery_contacts(email: str, phone: str) -> None:
    if email:
        set_app_setting(MANAGER_EMAIL_SETTING, email)
    else:
        delete_app_setting(MANAGER_EMAIL_SETTING)

    if phone:
        set_app_setting(MANAGER_PHONE_SETTING, phone)
    else:
        delete_app_setting(MANAGER_PHONE_SETTING)


def clear_manager_reset_code() -> None:
    delete_app_setting(MANAGER_RESET_CODE_HASH_SETTING)
    delete_app_setting(MANAGER_RESET_CODE_EXPIRY_SETTING)
    delete_app_setting(MANAGER_RESET_CODE_EMAIL_SETTING)
    delete_app_setting(MANAGER_RESET_CODE_TARGET_SETTING)
    delete_app_setting(MANAGER_RESET_CODE_CHANNEL_SETTING)


def active_manager_reset_request() -> dict[str, Any] | None:
    code_hash = app_setting(MANAGER_RESET_CODE_HASH_SETTING)
    expiry_value = app_setting(MANAGER_RESET_CODE_EXPIRY_SETTING)
    target_contact = (
        app_setting(MANAGER_RESET_CODE_TARGET_SETTING)
        or app_setting(MANAGER_RESET_CODE_EMAIL_SETTING)
    )
    channel = app_setting(MANAGER_RESET_CODE_CHANNEL_SETTING) or "email"
    expires_at = parse_stored_datetime(expiry_value)

    if (
        not code_hash
        or not target_contact
        or channel not in {"email", "phone"}
        or not expires_at
    ):
        clear_manager_reset_code()
        return None

    if now_ist() > expires_at:
        clear_manager_reset_code()
        return None

    return {
        "code_hash": code_hash,
        "contact": target_contact,
        "channel": channel,
        "expires_at": expires_at,
    }


def issue_manager_reset_code(
    target_contact: str,
    channel: str,
) -> str:
    code = "".join(
        secrets.choice("0123456789")
        for _ in range(MANAGER_RESET_CODE_LENGTH)
    )
    expires_at = now_ist() + MANAGER_RESET_CODE_TTL
    set_app_setting(
        MANAGER_RESET_CODE_HASH_SETTING,
        generate_password_hash(code),
    )
    set_app_setting(
        MANAGER_RESET_CODE_EXPIRY_SETTING,
        expires_at.isoformat(timespec="seconds"),
    )
    set_app_setting(MANAGER_RESET_CODE_TARGET_SETTING, target_contact)
    set_app_setting(MANAGER_RESET_CODE_CHANNEL_SETTING, channel)
    if channel == "email":
        set_app_setting(MANAGER_RESET_CODE_EMAIL_SETTING, target_contact)
    else:
        delete_app_setting(MANAGER_RESET_CODE_EMAIL_SETTING)
    return code


def verify_manager_reset_code(candidate: str) -> bool:
    state = active_manager_reset_request()
    if not state:
        return False
    return check_password_hash(state["code_hash"], candidate)


def deliver_manager_reset_code(
    target_contact: str,
    channel: str,
) -> str:
    code = issue_manager_reset_code(target_contact, channel)
    try:
        if channel == "email":
            if not manager_reset_email_delivery_ready():
                raise RuntimeError(
                    "Email delivery is not configured yet. "
                    "Set the SMTP environment values first."
                )
            send_manager_reset_code_email(target_contact, code)
        else:
            if not manager_reset_phone_delivery_ready():
                raise RuntimeError(
                    "SMS sending is not configured yet. "
                    "Set the Twilio environment values first."
                )
            send_manager_reset_code_phone(target_contact, code)
    except RuntimeError:
        clear_manager_reset_code()
        raise
    return code


def clear_manager_reset_verification() -> None:
    session.pop(MANAGER_RESET_VERIFIED_SESSION_KEY, None)


def active_manager_reset_verification() -> dict[str, Any] | None:
    state = session.get(MANAGER_RESET_VERIFIED_SESSION_KEY)
    if not isinstance(state, dict):
        clear_manager_reset_verification()
        return None

    verified_at = parse_stored_datetime(state.get("verified_at"))
    contact = (state.get("contact") or "").strip()
    channel = (state.get("channel") or "").strip()
    if (
        not verified_at
        or not contact
        or channel not in {"email", "phone"}
        or now_ist() - verified_at > MANAGER_RESET_VERIFIED_TTL
    ):
        clear_manager_reset_verification()
        return None

    return {
        "verified_at": verified_at,
        "contact": contact,
        "channel": channel,
    }


def mark_manager_reset_verified(
    target_contact: str,
    channel: str,
) -> None:
    session[MANAGER_RESET_VERIFIED_SESSION_KEY] = {
        "contact": target_contact,
        "channel": channel,
        "verified_at": iso_now(),
    }


def set_manager_reset_resend_cooldown() -> None:
    session[MANAGER_RESET_RESEND_SESSION_KEY] = (
        now_ist() + MANAGER_RESET_RESEND_COOLDOWN
    ).isoformat(timespec="seconds")


def manager_reset_resend_seconds_remaining() -> int:
    value = session.get(MANAGER_RESET_RESEND_SESSION_KEY)
    available_at = (
        parse_stored_datetime(value) if isinstance(value, str) else None
    )
    if not available_at:
        session.pop(MANAGER_RESET_RESEND_SESSION_KEY, None)
        return 0

    remaining = int((available_at - now_ist()).total_seconds())
    if remaining <= 0:
        session.pop(MANAGER_RESET_RESEND_SESSION_KEY, None)
        return 0
    return remaining


def clear_manager_reset_success() -> None:
    session.pop(MANAGER_RESET_SUCCESS_SESSION_KEY, None)


def manager_reset_success_active() -> bool:
    value = session.get(MANAGER_RESET_SUCCESS_SESSION_KEY)
    if not isinstance(value, str):
        clear_manager_reset_success()
        return False

    updated_at = parse_stored_datetime(value)
    if not updated_at or now_ist() - updated_at > timedelta(minutes=10):
        clear_manager_reset_success()
        return False
    return True


def mark_manager_reset_success() -> None:
    session[MANAGER_RESET_SUCCESS_SESSION_KEY] = iso_now()


def send_manager_reset_code_email(
    recipient_email: str,
    code: str,
) -> None:
    smtp_host = os.getenv("GODIGITAL_SMTP_HOST")
    from_email = os.getenv("GODIGITAL_SMTP_FROM_EMAIL")
    if not smtp_host or not from_email:
        raise RuntimeError(
            "Email sending is not configured yet. "
            "Set SMTP environment values first."
        )

    message = EmailMessage()
    message["Subject"] = "GoDigital manager password reset code"
    message["From"] = from_email
    message["To"] = recipient_email
    message.set_content(
        "Use this verification code to reset the "
        "GoDigital manager password:\n\n"
        f"{code}\n\n"
        "The code expires in 10 minutes. "
        "If you did not request this, you can ignore this email."
    )

    smtp_port = int(os.getenv("GODIGITAL_SMTP_PORT", "587"))
    smtp_username = os.getenv("GODIGITAL_SMTP_USERNAME") or ""
    smtp_password = os.getenv("GODIGITAL_SMTP_PASSWORD") or ""
    use_ssl = os.getenv("GODIGITAL_SMTP_USE_SSL", "0") == "1"
    use_tls = os.getenv("GODIGITAL_SMTP_USE_TLS", "1") == "1"

    try:
        if use_ssl:
            with smtplib.SMTP_SSL(
                smtp_host,
                smtp_port,
                timeout=20,
            ) as smtp:
                if smtp_username:
                    smtp.login(smtp_username, smtp_password)
                smtp.send_message(message)
        else:
            with smtplib.SMTP(
                smtp_host,
                smtp_port,
                timeout=20,
            ) as smtp:
                smtp.ehlo()
                if use_tls:
                    smtp.starttls()
                    smtp.ehlo()
                if smtp_username:
                    smtp.login(smtp_username, smtp_password)
                smtp.send_message(message)
    except (OSError, smtplib.SMTPException) as exc:
        app.logger.exception(
            "Failed to send manager reset email to %s",
            recipient_email,
        )
        raise RuntimeError(
            "Could not send the recovery email. Check the SMTP settings."
        ) from exc


def send_manager_reset_code_phone(
    recipient_phone: str,
    code: str,
) -> None:
    account_sid = os.getenv("GODIGITAL_TWILIO_ACCOUNT_SID")
    auth_token = os.getenv("GODIGITAL_TWILIO_AUTH_TOKEN")
    from_phone = os.getenv("GODIGITAL_TWILIO_FROM_PHONE")
    if not account_sid or not auth_token or not from_phone:
        raise RuntimeError(
            "SMS sending is not configured yet. "
            "Set the Twilio environment values first."
        )

    message_body = (
        "GoDigital manager password reset code: "
        f"{code}. This code expires in 10 minutes."
    )
    payload = urlencode(
        {
            "To": phone_for_sms_delivery(recipient_phone),
            "From": from_phone.strip(),
            "Body": message_body,
        }
    ).encode()
    auth_value = base64.b64encode(
        f"{account_sid}:{auth_token}".encode("utf-8")
    ).decode("ascii")
    request_obj = Request(
        (
            "https://api.twilio.com/2010-04-01/Accounts/"
            f"{account_sid}/Messages.json"
        ),
        data=payload,
        method="POST",
        headers={
            "Authorization": f"Basic {auth_value}",
            "Content-Type": "application/x-www-form-urlencoded",
        },
    )

    try:
        with urlopen(request_obj, timeout=20) as response:
            if response.status not in {200, 201}:
                raise RuntimeError("Unexpected SMS provider response.")
    except OSError as exc:
        app.logger.exception(
            "Failed to send manager reset SMS to %s",
            recipient_phone,
        )
        raise RuntimeError(
            "Could not send the recovery SMS. Check the Twilio settings."
        ) from exc


def manager_session_active() -> bool:
    timestamp = session.get(MANAGER_SESSION_KEY)
    if not isinstance(timestamp, str):
        return False

    authenticated_at = parse_stored_datetime(timestamp)
    if not authenticated_at:
        session.pop(MANAGER_SESSION_KEY, None)
        return False

    if now_ist() - authenticated_at > MANAGER_SESSION_DURATION:
        session.pop(MANAGER_SESSION_KEY, None)
        return False

    return True


def set_manager_session() -> None:
    session[MANAGER_SESSION_KEY] = iso_now()
    session["_csrf_token"] = secrets.token_urlsafe(32)
    session.permanent = True


def clear_manager_session() -> None:
    session.pop(MANAGER_SESSION_KEY, None)
    session["_csrf_token"] = secrets.token_urlsafe(32)


def verify_manager_password(candidate: str) -> bool:
    configured_hash = os.getenv("GODIGITAL_MANAGER_PASSWORD_HASH")
    if configured_hash:
        return check_password_hash(configured_hash, candidate)

    configured_plain = os.getenv("GODIGITAL_MANAGER_PASSWORD")
    if configured_plain:
        return hmac.compare_digest(candidate, configured_plain)

    stored_hash = app_setting(MANAGER_PASSWORD_SETTING)
    if not stored_hash:
        return False
    return check_password_hash(stored_hash, candidate)


def sanitize_next_target(value: str | None) -> str:
    candidate = (value or "").strip()
    if not candidate:
        return url_for("manager_page")

    parsed = urlsplit(candidate)
    if parsed.scheme or parsed.netloc:
        return url_for("manager_page")
    if not candidate.startswith("/") or candidate.startswith("//"):
        return url_for("manager_page")
    return candidate


def manager_access_context(
    next_target: str,
    flow: str = "unlock",
) -> dict[str, Any]:
    setup_mode = not manager_password_configured()
    active_flow = "setup" if setup_mode else "unlock"
    return {
        "page_title": "Manager Access",
        "next_target": next_target,
        "setup_mode": setup_mode,
        "access_flow": active_flow,
    }


def manager_access_redirect() -> Any:
    next_target = request.full_path if request.query_string else request.path
    return redirect(
        url_for(
            "manager_access_page",
            flow="unlock",
            next=sanitize_next_target(next_target),
        )
    )


def require_manager_access(view_func: Any) -> Any:
    @wraps(view_func)
    def wrapped(*args: Any, **kwargs: Any) -> Any:
        if manager_session_active():
            session[MANAGER_SESSION_KEY] = iso_now()
            return view_func(*args, **kwargs)
        return manager_access_redirect()

    return wrapped


@app.errorhandler(CSRFError)
def handle_csrf_error(error: CSRFError) -> Any:
    message = error.description or (
        "Security validation failed. Refresh the page and try again."
    )
    if request.path.startswith("/api/"):
        return jsonify({"success": False, "message": message}), 403
    flash(message, "error")
    return redirect(request.referrer or url_for("dashboard"))


@app.after_request
def apply_security_headers(response: Any) -> Any:
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    response.headers.setdefault(
        "Referrer-Policy",
        "strict-origin-when-cross-origin",
    )
    response.headers.setdefault("Cross-Origin-Opener-Policy", "same-origin")
    response.headers.setdefault("Cross-Origin-Resource-Policy", "same-origin")
    response.headers.setdefault("X-Permitted-Cross-Domain-Policies", "none")
    response.headers.setdefault(
        "Permissions-Policy",
        "camera=(self), microphone=(), geolocation=()",
    )
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; "
        "img-src 'self' data: blob:; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline'; "
        "font-src 'self' data:; "
        "connect-src 'self'; "
        "media-src 'self' blob:; "
        "object-src 'none'; "
        "base-uri 'self'; "
        "form-action 'self'; "
        "frame-ancestors 'self';",
    )
    if request.is_secure or app.config["SESSION_COOKIE_SECURE"]:
        response.headers.setdefault(
            "Strict-Transport-Security",
            "max-age=31536000; includeSubDomains",
        )
    if (
        request.method == "GET"
        and request.endpoint != "static"
        and response.mimetype == "text/html"
    ):
        response.headers["Cache-Control"] = "private, no-store, max-age=0"
        response.headers["Pragma"] = "no-cache"
    return response


def initials(value: str | None) -> str:
    parts = [part for part in (value or "").split() if part]
    if not parts:
        return "GD"
    if len(parts) == 1:
        return parts[0][:2].upper()
    return f"{parts[0][0]}{parts[-1][0]}".upper()


def employee_view(row: sqlite3.Row) -> dict[str, Any]:
    department_value = normalize_department(row["department"])
    work_mode_value = normalize_work_mode(row["work_mode"])
    return {
        "id": row["id"],
        "employee_id": row["employee_id"],
        "name": row["name"] or "Unknown Employee",
        "phone": row["phone"] or "--",
        "phone_value": row["phone"] or "",
        "department": department_value,
        "department_value": department_value,
        "work_mode": work_mode_value,
        "work_mode_value": work_mode_value,
        "photo_url": photo_url(row["photo_path"]),
        "initials": initials(row["name"]),
        "created_at_label": format_datetime_label(row["created_at"]),
    }


def attendance_view(row: sqlite3.Row) -> dict[str, Any]:
    business_date = parse_business_date(row["date"])
    department_value = normalize_department(row["department"])
    row_keys = set(row.keys())
    attendance_work_mode = (
        row["work_mode"] if "work_mode" in row_keys else None
    )
    employee_work_mode = (
        row["employee_work_mode"] if "employee_work_mode" in row_keys else None
    )
    work_mode_value = normalize_work_mode(
        attendance_work_mode or employee_work_mode
    )
    live_hours = compute_live_hours(
        row["check_in"],
        row["check_out"],
        business_date,
    )
    saved_hours = float(row["total_hours"] or 0.0)
    total_hours = saved_hours if row["check_out"] else live_hours
    if row["check_out"] and saved_hours == 0:
        total_hours = compute_total_hours(
            row["check_in"],
            row["check_out"],
            business_date,
        )

    attendance_type = row["attendance_type"] or (
        "In Progress"
        if row["check_in"] and not row["check_out"]
        else "Pending"
    )
    status = row["status"] or "Pending"

    return {
        "id": row["id"],
        "employee_id": row["employee_id"],
        "name": row["name"] or row["employee_id"],
        "phone": row["phone"] or "--",
        "department": department_value,
        "work_mode": work_mode_value,
        "work_mode_tone": badge_tone_for_work_mode(work_mode_value),
        "photo_url": photo_url(row["photo_path"]),
        "initials": initials(row["name"]),
        "date": business_date.isoformat(),
        "date_label": business_date.strftime("%d %b %Y"),
        "check_in": row["check_in"],
        "check_out": row["check_out"],
        "check_in_label": format_time_label(row["check_in"], business_date),
        "check_out_label": format_time_label(row["check_out"], business_date),
        "total_hours": round(total_hours, 2),
        "hours_label": hours_to_label(total_hours) if total_hours else (
            "In progress" if row["check_in"] and not row["check_out"] else "--"
        ),
        "status": status,
        "status_tone": badge_tone_for_status(status),
        "attendance_type": attendance_type,
        "attendance_type_tone": badge_tone_for_attendance_type(
            attendance_type
        ),
        "check_in_input": time_input_value(row["check_in"], business_date),
        "check_out_input": time_input_value(row["check_out"], business_date),
    }


def employee_rows() -> list[sqlite3.Row]:
    return query_all(
        """
        SELECT *
        FROM employees
        ORDER BY name COLLATE NOCASE ASC
        """
    )


def employee_by_id(employee_pk: int) -> sqlite3.Row | None:
    return query_one("SELECT * FROM employees WHERE id = ?", (employee_pk,))


def employee_by_code(employee_code: str) -> sqlite3.Row | None:
    return query_one(
        "SELECT * FROM employees WHERE employee_id = ?",
        (employee_code,),
    )


def employee_by_enrollment_details(
    name: str,
    phone: str,
) -> sqlite3.Row | None:
    target_name = normalize_person_name(name)
    target_phone = normalize_phone_lookup(phone)
    if not target_phone:
        return None

    for row in employee_rows():
        row_phone = normalize_phone_lookup(row["phone"])
        row_name = normalize_person_name(row["name"])
        if row_phone == target_phone:
            return row
        if row_phone and row_name == target_name and row_phone == target_phone:
            return row
    return None


def today_record_for_employee(employee_code: str) -> sqlite3.Row | None:
    return query_one(
        """
        SELECT *
        FROM attendance
        WHERE employee_id = ? AND date = ?
        """,
        (employee_code, today_ist().isoformat()),
    )


def attendance_by_id(record_id: int) -> sqlite3.Row | None:
    return query_one(
        """
        SELECT
            a.*,
            e.name,
            e.phone,
            e.department,
            e.work_mode AS employee_work_mode,
            e.photo_path
        FROM attendance a
        LEFT JOIN employees e ON e.employee_id = a.employee_id
        WHERE a.id = ?
        """,
        (record_id,),
    )


def generate_employee_code() -> str:
    row = query_one("SELECT MAX(id) AS max_id FROM employees")
    next_id = int(row["max_id"] or 0) + 1
    return f"GD{next_id:03d}"


def monthly_summary_rows(month_ref: str) -> list[dict[str, Any]]:
    month_value = (
        month_ref
        if re.fullmatch(r"\d{4}-\d{2}", month_ref or "")
        else today_ist().strftime("%Y-%m")
    )
    rows = query_all(
        """
        SELECT
            e.id,
            e.employee_id,
            e.name,
            e.department,
            COUNT(a.id) AS present_days,
            SUM(
                CASE WHEN a.status = 'On Time' THEN 1 ELSE 0 END
            ) AS on_time_days,
            SUM(CASE WHEN a.status = 'Late' THEN 1 ELSE 0 END) AS late_days,
            SUM(
                CASE WHEN a.attendance_type = 'Half Day' THEN 1 ELSE 0 END
            ) AS half_days,
            SUM(
                CASE WHEN a.attendance_type = 'Full Day' THEN 1 ELSE 0 END
            ) AS full_days,
            ROUND(COALESCE(SUM(a.total_hours), 0), 2) AS total_hours
        FROM employees e
        LEFT JOIN attendance a
            ON a.employee_id = e.employee_id
            AND substr(a.date, 1, 7) = ?
        GROUP BY e.id, e.employee_id, e.name, e.department
        ORDER BY e.name COLLATE NOCASE ASC
        """,
        (month_value,),
    )

    summary: list[dict[str, Any]] = []
    for row in rows:
        summary.append(
            {
                "id": row["id"],
                "employee_id": row["employee_id"],
                "name": row["name"] or row["employee_id"],
                "department": normalize_department(row["department"]),
                "present_days": row["present_days"] or 0,
                "on_time_days": row["on_time_days"] or 0,
                "late_days": row["late_days"] or 0,
                "half_days": row["half_days"] or 0,
                "full_days": row["full_days"] or 0,
                "total_hours": float(row["total_hours"] or 0),
                "hours_label": hours_to_label(float(row["total_hours"] or 0)),
            }
        )
    return summary


def history_filters_from_request(
    default_days: int = 30,
) -> dict[str, str]:
    today_value = today_ist()
    start_default = (
        today_value - timedelta(days=default_days - 1)
    ).isoformat()
    return {
        "start_date": (
            request.args.get("start_date") or start_default
        ).strip(),
        "end_date": (
            request.args.get("end_date") or today_value.isoformat()
        ).strip(),
        "employee_id": (request.args.get("employee_id") or "").strip(),
    }


def history_rows(filters: dict[str, str]) -> list[dict[str, Any]]:
    clauses = ["1 = 1"]
    params: list[Any] = []

    if filters["start_date"]:
        clauses.append("a.date >= ?")
        params.append(filters["start_date"])
    if filters["end_date"]:
        clauses.append("a.date <= ?")
        params.append(filters["end_date"])
    if filters["employee_id"]:
        clauses.append("a.employee_id = ?")
        params.append(filters["employee_id"])

    rows = query_all(
        f"""
        SELECT
            a.*,
            e.name,
            e.phone,
            e.department,
            e.work_mode AS employee_work_mode,
            e.photo_path
        FROM attendance a
        LEFT JOIN employees e ON e.employee_id = a.employee_id
        WHERE {' AND '.join(clauses)}
        ORDER BY
            a.date DESC,
            COALESCE(a.check_in, a.check_out) DESC,
            e.name COLLATE NOCASE ASC
        """,
        tuple(params),
    )
    return [attendance_view(row) for row in rows]


def build_today_bundle() -> dict[str, Any]:
    rows = query_all(
        """
        SELECT
            a.*,
            e.name,
            e.phone,
            e.department,
            e.work_mode AS employee_work_mode,
            e.photo_path
        FROM attendance a
        LEFT JOIN employees e ON e.employee_id = a.employee_id
        WHERE a.date = ?
        ORDER BY
            COALESCE(a.check_in, a.check_out) DESC,
            e.name COLLATE NOCASE ASC
        """,
        (today_ist().isoformat(),),
    )
    records = [attendance_view(row) for row in rows]

    return {
        "date_label": current_date_label(),
        "short_date_label": short_date_label(),
        "present_today": sum(1 for row in records if row["check_in"]),
        "on_time": sum(1 for row in records if row["status"] == "On Time"),
        "late": sum(1 for row in records if row["status"] == "Late"),
        "half_day": sum(
            1 for row in records if row["attendance_type"] == "Half Day"
        ),
        "records": records,
        "last_updated": current_system_label(),
    }


def export_excel_rows(filters: dict[str, str]) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    rows = history_rows(filters)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Attendance"

    headers = [
        "Employee ID",
        "Name",
        "Phone",
        "Department",
        "Work Mode",
        "Date",
        "Check In",
        "Check Out",
        "Hours",
        "Status",
        "Attendance Type",
    ]
    worksheet.append(headers)

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        worksheet.append(
            [
                row["employee_id"],
                row["name"],
                row["phone"],
                row["department"],
                row["work_mode"],
                parse_business_date(row["date"]),
                row["check_in_label"],
                row["check_out_label"],
                row["hours_label"],
                row["status"],
                row["attendance_type"],
            ]
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet["F"][1:]:
        cell.number_format = "DD-MMM-YYYY"

    column_widths = {
        "A": 14,
        "B": 24,
        "C": 16,
        "D": 18,
        "E": 14,
        "F": 14,
        "G": 12,
        "H": 12,
        "I": 12,
        "J": 14,
        "K": 18,
    }
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


class RestrictedFaceEncodingUnpickler(pickle.Unpickler):
    SAFE_GLOBALS = {
        ("numpy", "dtype"): np.dtype,
        ("numpy", "ndarray"): np.ndarray,
        (
            "numpy._core.multiarray",
            "_reconstruct",
        ): np.core.multiarray._reconstruct,
        (
            "numpy.core.multiarray",
            "_reconstruct",
        ): np.core.multiarray._reconstruct,
        ("builtins", "tuple"): tuple,
    }

    def find_class(self, module: str, name: str) -> Any:
        allowed = self.SAFE_GLOBALS.get((module, name))
        if allowed is None:
            raise pickle.UnpicklingError(
                "Unsupported serialized face encoding."
            )
        return allowed


def serialize_face_encoding(face_encoding: np.ndarray) -> bytes:
    buffer = BytesIO()
    np.save(
        buffer,
        np.asarray(face_encoding, dtype=np.float64),
        allow_pickle=False,
    )
    return buffer.getvalue()


def deserialize_face_encoding(value: bytes | None) -> np.ndarray | None:
    if not value:
        return None

    try:
        encoding = np.load(BytesIO(value), allow_pickle=False)
    except Exception:
        try:
            legacy_value = RestrictedFaceEncodingUnpickler(
                BytesIO(value)
            ).load()
        except Exception:
            return None
        encoding = np.asarray(legacy_value, dtype=np.float64)

    encoding = np.asarray(encoding, dtype=np.float64)
    if encoding.ndim != 1 or encoding.size != 128:
        return None
    return encoding


def decode_data_url(image_data: str) -> tuple[bytes, np.ndarray]:
    if not image_data:
        raise ValueError("Image data is required.")
    if not isinstance(image_data, str):
        raise ValueError("Captured image payload is invalid.")
    if len(image_data) > MAX_IMAGE_DATA_LENGTH:
        raise ValueError("Captured image is too large. Please try again.")

    prefix = image_data.split(",", 1)[0] if "," in image_data else ""
    mime_match = re.fullmatch(
        r"data:(image/[a-z0-9.+-]+);base64",
        prefix,
        re.IGNORECASE,
    )
    if (
        mime_match
        and mime_match.group(1).lower() not in ALLOWED_IMAGE_MIME_TYPES
    ):
        raise ValueError(
            "Unsupported image format. Please capture the photo again."
        )

    encoded = image_data.split(",", 1)[1] if "," in image_data else image_data
    try:
        raw_bytes = base64.b64decode(encoded, validate=True)
    except Exception as exc:
        raise ValueError("Could not decode the captured image.") from exc
    if len(raw_bytes) > MAX_FACE_IMAGE_BYTES:
        raise ValueError(
            "Captured image is too large. Please retake the photo."
        )

    try:
        with Image.open(BytesIO(raw_bytes)) as pil_image:
            if pil_image.format not in {"JPEG", "PNG", "WEBP"}:
                raise ValueError(
                    "Unsupported image format. Please capture the photo again."
                )
            rgb_image = np.array(pil_image.convert("RGB"))
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(
            "Captured image could not be processed. Please try again."
        ) from exc

    if rgb_image.ndim != 3:
        raise ValueError(
            "Captured image could not be processed. Please try again."
        )
    if rgb_image.shape[0] < 160 or rgb_image.shape[1] < 160:
        raise ValueError(
            "Captured image is too small. Move closer and try again."
        )

    if cv2 is not None:
        bgr_image = cv2.cvtColor(rgb_image, cv2.COLOR_RGB2BGR)
        height, width = bgr_image.shape[:2]
        longest_side = max(height, width)
        if longest_side > 1280:
            scale = 1280 / float(longest_side)
            bgr_image = cv2.resize(
                bgr_image,
                None,
                fx=scale,
                fy=scale,
                interpolation=cv2.INTER_AREA,
            )
        rgb_image = cv2.cvtColor(bgr_image, cv2.COLOR_BGR2RGB)

    return raw_bytes, rgb_image


def require_face_stack() -> None:
    if cv2 is None:
        raise RuntimeError(
            "OpenCV is unavailable in this environment: "
            f"{CV2_IMPORT_ERROR}. {face_stack_runtime_hint()}"
        )
    if face_recognition is None:
        raise RuntimeError(
            "face_recognition is unavailable in this environment: "
            f"{FACE_IMPORT_ERROR}. {face_stack_runtime_hint()}"
        )


def extract_single_face_encoding(image_data: str) -> tuple[bytes, np.ndarray]:
    require_face_stack()
    raw_bytes, rgb_image = decode_data_url(image_data)

    face_locations = face_recognition.face_locations(rgb_image, model="hog")
    if not face_locations:
        raise ValueError("No face detected. Center your face and try again.")
    if len(face_locations) > 1:
        raise ValueError("Multiple faces detected. Show one face only.")

    encodings = face_recognition.face_encodings(
        rgb_image,
        known_face_locations=face_locations,
    )
    if not encodings:
        raise ValueError(
            "A face was detected, but encoding could not be created."
        )

    return raw_bytes, encodings[0]


def save_employee_photo(image_bytes: bytes, employee_code: str) -> str:
    ensure_upload_dir()
    filename = (
        f"{slugify(employee_code)}-"
        f"{now_ist().strftime('%Y%m%d%H%M%S')}.jpg"
    )
    save_path = UPLOAD_DIR / filename

    with Image.open(BytesIO(image_bytes)) as image:
        image = image.convert("RGB")
        image.thumbnail((900, 900))
        image.save(save_path, format="JPEG", quality=90)

    return f"uploads/employees/{filename}"


def known_face_rows() -> tuple[list[np.ndarray], list[sqlite3.Row]]:
    rows = query_all(
        """
        SELECT *
        FROM employees
        WHERE face_encoding IS NOT NULL
        ORDER BY id ASC
        """
    )

    known_encodings: list[np.ndarray] = []
    known_rows: list[sqlite3.Row] = []

    for row in rows:
        encoding = deserialize_face_encoding(row["face_encoding"])
        if encoding is None:
            continue
        known_encodings.append(encoding)
        known_rows.append(row)

    return known_encodings, known_rows


def known_face_count() -> int:
    row = query_one(
        """
        SELECT COUNT(*) AS total
        FROM employees
        WHERE face_encoding IS NOT NULL
        """
    )
    return int(row["total"] or 0) if row else 0


def match_employee_from_encoding(
    face_encoding: np.ndarray,
) -> sqlite3.Row | None:
    require_face_stack()
    known_encodings, rows = known_face_rows()
    if not known_encodings:
        return None

    matches = face_recognition.compare_faces(
        known_encodings,
        face_encoding,
        tolerance=0.5,
    )
    distances = face_recognition.face_distance(known_encodings, face_encoding)
    best_index = int(np.argmin(distances))
    if matches[best_index]:
        return rows[best_index]
    return None


def validate_attendance_times(
    check_in_value: str | None,
    check_out_value: str | None,
    business_date: date,
) -> str | None:
    check_in_dt = parse_stored_datetime(check_in_value, business_date)
    check_out_dt = parse_stored_datetime(check_out_value, business_date)

    if not check_in_dt:
        return "Check-in time is required."
    if check_out_dt and check_out_dt <= check_in_dt:
        return "Check-out must be later than check-in."
    return None


def employee_exists(employee_code: str) -> bool:
    return employee_by_code(employee_code) is not None


def upsert_attendance_record(
    employee_code: str,
    business_date: date,
    check_in_value: str | None,
    check_out_value: str | None,
    work_mode_value: str | None = None,
    status_value: str | None = None,
    attendance_type_value: str | None = None,
) -> sqlite3.Row:
    record = query_one(
        """
        SELECT *
        FROM attendance
        WHERE employee_id = ? AND date = ?
        """,
        (employee_code, business_date.isoformat()),
    )

    check_in_dt = parse_stored_datetime(check_in_value, business_date)
    total_hours = compute_total_hours(
        check_in_value,
        check_out_value,
        business_date,
    )
    resolved_work_mode = work_mode_value
    if resolved_work_mode is None and record and record["work_mode"]:
        resolved_work_mode = normalize_work_mode(record["work_mode"])
    if resolved_work_mode is None:
        employee = employee_by_code(employee_code)
        resolved_work_mode = normalize_work_mode(
            employee["work_mode"] if employee else None
        )
    status = status_value or status_from_checkin(check_in_dt)
    attendance_type = attendance_type_value or attendance_type_from_hours(
        total_hours,
        bool(check_out_value),
    )

    if record:
        execute_db(
            """
            UPDATE attendance
            SET check_in = ?,
                check_out = ?,
                work_mode = ?,
                total_hours = ?,
                status = ?,
                attendance_type = ?
            WHERE id = ?
            """,
            (
                check_in_value,
                check_out_value,
                resolved_work_mode,
                total_hours,
                status,
                attendance_type,
                record["id"],
            ),
        )
    else:
        execute_db(
            """
            INSERT INTO attendance (
                employee_id,
                date,
                check_in,
                check_out,
                work_mode,
                total_hours,
                status,
                attendance_type
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                employee_code,
                business_date.isoformat(),
                check_in_value,
                check_out_value,
                resolved_work_mode,
                total_hours,
                status,
                attendance_type,
            ),
        )

    refreshed = query_one(
        """
        SELECT
            a.*,
            e.name,
            e.phone,
            e.department,
            e.work_mode AS employee_work_mode,
            e.photo_path
        FROM attendance a
        LEFT JOIN employees e ON e.employee_id = a.employee_id
        WHERE a.employee_id = ? AND a.date = ?
        """,
        (employee_code, business_date.isoformat()),
    )
    assert refreshed is not None
    return refreshed


def process_scan_for_employee(
    employee_row: sqlite3.Row,
    work_mode: str | None = None,
) -> dict[str, Any]:
    employee_code = employee_row["employee_id"]
    employee_name = employee_row["name"]
    business_date = today_ist()
    current_record = today_record_for_employee(employee_code)
    current_timestamp = iso_now()
    selected_work_mode = normalize_work_mode(
        work_mode
        or (
            current_record["work_mode"]
            if current_record and current_record["work_mode"]
            else employee_row["work_mode"]
        )
    )

    if not current_record or not current_record["check_in"]:
        updated = upsert_attendance_record(
            employee_code,
            business_date,
            current_timestamp,
            None,
            work_mode_value=selected_work_mode,
            status_value=status_from_checkin(
                parse_stored_datetime(current_timestamp, business_date)
            ),
            attendance_type_value="In Progress",
        )
        action = "check_in"
        message = (
            f"{employee_name} checked in at "
            f"{format_time_label(current_timestamp, business_date)}."
        )
    elif current_record["check_in"] and not current_record["check_out"]:
        total_hours = compute_total_hours(
            current_record["check_in"],
            current_timestamp,
            business_date,
        )
        updated = upsert_attendance_record(
            employee_code,
            business_date,
            current_record["check_in"],
            current_timestamp,
            work_mode_value=selected_work_mode,
            status_value=(
                current_record["status"]
                or status_from_checkin(
                    parse_stored_datetime(
                        current_record["check_in"],
                        business_date,
                    )
                )
            ),
            attendance_type_value=attendance_type_from_hours(
                total_hours,
                True,
            ),
        )
        action = "check_out"
        message = (
            f"{employee_name} checked out at "
            f"{format_time_label(current_timestamp, business_date)}."
        )
    else:
        updated = attendance_by_id(current_record["id"])
        assert updated is not None
        action = "already_marked"
        message = (
            "You have already recorded attendance today "
            f"for {current_date_label(business_date)}."
        )

    return {
        "success": True,
        "recognized": True,
        "action": action,
        "message": message,
        "employee": employee_view(employee_row),
        "attendance": attendance_view(updated),
    }


def set_pending_scan_employee(employee_code: str) -> None:
    session[PENDING_SCAN_SESSION_KEY] = {
        "employee_id": employee_code,
        "created_at": iso_now(),
    }


def clear_pending_scan_employee() -> None:
    session.pop(PENDING_SCAN_SESSION_KEY, None)


def pending_scan_employee_code() -> str | None:
    pending = session.get(PENDING_SCAN_SESSION_KEY)
    if not isinstance(pending, dict):
        clear_pending_scan_employee()
        return None

    employee_code = (pending.get("employee_id") or "").strip()
    created_at = parse_stored_datetime(pending.get("created_at"))
    if (
        not employee_code
        or created_at is None
        or now_ist() - created_at > PENDING_SCAN_TTL
    ):
        clear_pending_scan_employee()
        return None

    return employee_code


@app.context_processor
def inject_globals() -> dict[str, Any]:
    return {
        "system_time_label": current_system_label(),
        "current_year": today_ist().year,
        "department_options": DEPARTMENT_OPTIONS,
        "work_mode_options": WORK_MODE_OPTIONS,
        "static_asset_url": static_asset_url,
        "csrf_token": csrf_token,
        "manager_session_active": manager_session_active(),
        "manager_access_configured": manager_password_configured(),
    }


@app.route("/home")
def home_redirect() -> Any:
    return redirect(url_for("dashboard"))


@app.route("/")
def dashboard() -> str:
    today_bundle = build_today_bundle()
    return render_template(
        "home/index.html",
        today_bundle=today_bundle,
        page_title="GoDigital Smart Attendance",
        hero_date_label=current_date_label(),
        hero_date_short=short_date_label(),
    )


@app.route("/scan")
def scan_page() -> str:
    return render_template(
        "attendance/face-scan.html",
        page_title="Face Scan",
        today_label=current_date_label(),
    )


@app.route("/enroll")
def enroll_page() -> str:
    return render_template(
        "attendance/enroll-face.html",
        page_title="Enroll Employee",
    )


@app.route("/history")
def history_page() -> str:
    filters = history_filters_from_request()
    records = history_rows(filters)
    employees = [employee_view(row) for row in employee_rows()]
    total_hours = round(sum(row["total_hours"] for row in records), 2)
    return render_template(
        "attendance/history.html",
        page_title="Attendance History",
        filters=filters,
        records=records,
        employees=employees,
        summary={
            "records": len(records),
            "hours": total_hours,
            "on_time": sum(1 for row in records if row["status"] == "On Time"),
            "late": sum(1 for row in records if row["status"] == "Late"),
        },
    )


@app.route("/manager")
@require_manager_access
def manager_page() -> str:
    filters = history_filters_from_request()
    employees = [employee_view(row) for row in employee_rows()]
    records = history_rows(filters)
    month_ref = (
        request.args.get("month") or today_ist().strftime("%Y-%m")
    ).strip()
    monthly_summary = monthly_summary_rows(month_ref)
    today_bundle = build_today_bundle()
    return render_template(
        "dashboard/admin.html",
        page_title="Manager Panel",
        employees=employees,
        records=records,
        filters=filters,
        month_ref=month_ref,
        monthly_summary=monthly_summary,
        today_bundle=today_bundle,
    )


@app.route("/manager/access")
def manager_access_page() -> str:
    if manager_session_active():
        return redirect(sanitize_next_target(request.args.get("next")))

    next_target = sanitize_next_target(request.args.get("next"))
    clear_manager_reset_code()
    clear_manager_reset_verification()
    clear_manager_reset_success()
    session.pop(MANAGER_RESET_RESEND_SESSION_KEY, None)
    return render_template(
        "manager/access.html",
        **manager_access_context(next_target),
    )


@app.post("/manager/access")
def manager_access_submit() -> Any:
    next_target = sanitize_next_target(request.form.get("next"))
    retry_after = consume_rate_limit(
        "manager-access",
        max_requests=8,
        window_seconds=600,
    )
    if retry_after:
        flash(
            "Too many manager unlock attempts. "
            f"Try again in about {retry_after} seconds.",
            "error",
        )
        return redirect(url_for("manager_access_page", next=next_target))

    password = request.form.get("password") or ""
    setup_mode = not manager_password_configured()

    if setup_mode:
        confirm_password = request.form.get("confirm_password") or ""
        if len(password) < 10:
            flash(
                "Manager password must be at least 10 characters long.",
                "error",
            )
            return render_template(
                "manager/access.html",
                **manager_access_context(next_target, "setup"),
            )
        if password != confirm_password:
            flash(
                "Manager passwords did not match. Please try again.",
                "error",
            )
            return render_template(
                "manager/access.html",
                **manager_access_context(next_target, "setup"),
            )

        set_app_setting(
            MANAGER_PASSWORD_SETTING,
            generate_password_hash(password),
        )
        clear_manager_reset_code()
        clear_manager_reset_verification()
        session.pop(MANAGER_RESET_RESEND_SESSION_KEY, None)
        clear_manager_reset_success()
        set_manager_session()
        flash(
            "Manager password created. "
            "Manager controls are now protected.",
            "success",
        )
        return redirect(next_target)

    if not verify_manager_password(password):
        flash("Incorrect manager password.", "error")
        return render_template(
            "manager/access.html",
            **manager_access_context(next_target, "unlock"),
        )

    clear_manager_reset_success()
    clear_manager_reset_verification()
    session.pop(MANAGER_RESET_RESEND_SESSION_KEY, None)
    set_manager_session()
    flash("Manager controls unlocked.", "success")
    return redirect(next_target)


@app.post("/manager/password/forgot")
def manager_password_forgot() -> Any:
    next_target = sanitize_next_target(request.form.get("next"))
    clear_manager_reset_code()
    clear_manager_reset_verification()
    session.pop(MANAGER_RESET_RESEND_SESSION_KEY, None)
    flash(MANAGER_PASSWORD_ONLY_MESSAGE, "warning")
    return redirect(url_for("manager_access_page", next=next_target))


@app.post("/manager/password/reset")
def manager_password_reset() -> Any:
    next_target = sanitize_next_target(request.form.get("next"))
    clear_manager_reset_code()
    clear_manager_reset_verification()
    session.pop(MANAGER_RESET_RESEND_SESSION_KEY, None)
    flash(MANAGER_PASSWORD_ONLY_MESSAGE, "warning")
    return redirect(url_for("manager_access_page", next=next_target))


@app.post("/manager/password/verify")
def manager_password_verify() -> Any:
    next_target = sanitize_next_target(request.form.get("next"))
    clear_manager_reset_code()
    session.pop(MANAGER_RESET_RESEND_SESSION_KEY, None)
    clear_manager_reset_verification()
    flash(MANAGER_PASSWORD_ONLY_MESSAGE, "warning")
    return redirect(url_for("manager_access_page", next=next_target))


@app.post("/manager/password/resend")
def manager_password_resend() -> Any:
    next_target = sanitize_next_target(request.form.get("next"))
    clear_manager_reset_code()
    clear_manager_reset_verification()
    session.pop(MANAGER_RESET_RESEND_SESSION_KEY, None)
    flash(MANAGER_PASSWORD_ONLY_MESSAGE, "warning")
    return redirect(url_for("manager_access_page", next=next_target))


@app.post("/manager/logout")
def manager_logout() -> Any:
    clear_manager_session()
    flash("Manager controls locked.", "success")
    return redirect(url_for("dashboard"))


@app.post("/manager/recovery")
@require_manager_access
def update_manager_recovery_contacts() -> Any:
    clear_manager_reset_code()
    clear_manager_reset_verification()
    session.pop(MANAGER_RESET_RESEND_SESSION_KEY, None)
    flash(MANAGER_RECOVERY_DISABLED_MESSAGE, "warning")
    return redirect(url_for("manager_page"))


@app.post("/manager/employees/<int:employee_pk>/update")
@require_manager_access
def update_employee(employee_pk: int) -> Any:
    employee = employee_by_id(employee_pk)
    if not employee:
        flash("Employee not found.", "error")
        return redirect(url_for("manager_page"))

    name, name_error = validate_employee_name(request.form.get("name"))
    if name_error:
        flash(name_error, "error")
        return redirect(url_for("manager_page"))

    phone, phone_error = validate_phone_number(
        request.form.get("phone"),
        required=False,
    )
    if phone_error:
        flash(phone_error, "error")
        return redirect(url_for("manager_page"))

    department = normalize_department(request.form.get("department"))
    work_mode = normalize_work_mode(request.form.get("work_mode"))

    execute_db(
        """
        UPDATE employees
        SET name = ?, phone = ?, department = ?, work_mode = ?
        WHERE id = ?
        """,
        (name, phone, department, work_mode, employee_pk),
    )
    flash(f"{name} was updated successfully.", "success")
    return redirect(url_for("manager_page"))


@app.post("/manager/employees/<int:employee_pk>/delete")
@require_manager_access
def delete_employee(employee_pk: int) -> Any:
    employee = employee_by_id(employee_pk)
    if not employee:
        flash("Employee not found.", "error")
        return redirect(url_for("manager_page"))

    execute_db(
        "DELETE FROM attendance WHERE employee_id = ?",
        (employee["employee_id"],),
    )
    execute_db("DELETE FROM employees WHERE id = ?", (employee_pk,))
    flash(f"{employee['name']} was deleted.", "success")
    return redirect(url_for("manager_page"))


@app.post("/manager/attendance/add")
@require_manager_access
def add_attendance_record() -> Any:
    employee_code = (request.form.get("employee_id") or "").strip()
    business_date = parse_business_date(request.form.get("date"))
    work_mode_value = normalize_optional_work_mode(
        request.form.get("work_mode")
    )
    check_in_value = build_iso_timestamp(
        business_date,
        request.form.get("check_in"),
    )
    check_out_value = build_iso_timestamp(
        business_date,
        request.form.get("check_out"),
    )

    if not employee_code or not check_in_value:
        flash("Employee and check-in time are required.", "error")
        return redirect(url_for("manager_page"))
    if not employee_exists(employee_code):
        flash("Select a valid employee before saving attendance.", "error")
        return redirect(url_for("manager_page"))

    time_error = validate_attendance_times(
        check_in_value,
        check_out_value,
        business_date,
    )
    if time_error:
        flash(time_error, "error")
        return redirect(url_for("manager_page"))

    status_value = normalize_status_input(request.form.get("status"))
    attendance_type_value = normalize_attendance_type_input(
        request.form.get("attendance_type")
    )
    if (request.form.get("status") or "").strip() and status_value is None:
        flash("Status value is invalid.", "error")
        return redirect(url_for("manager_page"))
    if (
        (request.form.get("attendance_type") or "").strip()
        and attendance_type_value is None
    ):
        flash("Attendance type value is invalid.", "error")
        return redirect(url_for("manager_page"))

    upsert_attendance_record(
        employee_code,
        business_date,
        check_in_value,
        check_out_value,
        work_mode_value=work_mode_value,
        status_value=status_value,
        attendance_type_value=attendance_type_value,
    )
    flash("Attendance record saved.", "success")
    return redirect(url_for("manager_page"))


@app.post("/manager/attendance/<int:record_id>/update")
@require_manager_access
def update_attendance_record(record_id: int) -> Any:
    record = attendance_by_id(record_id)
    if not record:
        flash("Attendance record not found.", "error")
        return redirect(url_for("manager_page"))

    business_date = parse_business_date(request.form.get("date"))
    work_mode_value = normalize_optional_work_mode(
        request.form.get("work_mode")
    )
    check_in_value = build_iso_timestamp(
        business_date,
        request.form.get("check_in"),
    )
    check_out_value = build_iso_timestamp(
        business_date,
        request.form.get("check_out"),
    )

    if not check_in_value:
        flash("Check-in time is required.", "error")
        return redirect(url_for("manager_page"))

    time_error = validate_attendance_times(
        check_in_value,
        check_out_value,
        business_date,
    )
    if time_error:
        flash(time_error, "error")
        return redirect(url_for("manager_page"))

    status_value = normalize_status_input(request.form.get("status"))
    attendance_type_value = normalize_attendance_type_input(
        request.form.get("attendance_type")
    )
    if (request.form.get("status") or "").strip() and status_value is None:
        flash("Status value is invalid.", "error")
        return redirect(url_for("manager_page"))
    if (
        (request.form.get("attendance_type") or "").strip()
        and attendance_type_value is None
    ):
        flash("Attendance type value is invalid.", "error")
        return redirect(url_for("manager_page"))

    upsert_attendance_record(
        record["employee_id"],
        business_date,
        check_in_value,
        check_out_value,
        work_mode_value=work_mode_value,
        status_value=status_value,
        attendance_type_value=attendance_type_value,
    )
    flash("Attendance record updated.", "success")
    return redirect(url_for("manager_page"))


@app.post("/manager/attendance/<int:record_id>/delete")
@require_manager_access
def delete_attendance_record(record_id: int) -> Any:
    execute_db("DELETE FROM attendance WHERE id = ?", (record_id,))
    flash("Attendance record deleted.", "success")
    return redirect(url_for("manager_page"))


@app.post("/api/scan")
@csrf.exempt
def api_scan() -> Any:
    retry_after = consume_rate_limit(
        "scan",
        max_requests=18,
        window_seconds=60,
    )
    if retry_after:
        return rate_limited_json_response(
            "Too many scan attempts. "
            "Please wait a moment before trying again.",
            retry_after,
        )

    payload = request.get_json(silent=True) or {}
    employee_code = (payload.get("employee_id") or "").strip()
    selected_work_mode = normalize_optional_work_mode(
        payload.get("work_mode")
    )

    if employee_code or selected_work_mode is not None:
        pending_employee_code = pending_scan_employee_code()
        if not employee_code or selected_work_mode is None:
            clear_pending_scan_employee()
            return jsonify(
                {
                    "success": False,
                    "recognized": False,
                    "code": "work_mode_required",
                    "message": (
                        "Choose today's work mode before recording "
                        "attendance."
                    ),
                }
            )
        if pending_employee_code != employee_code:
            clear_pending_scan_employee()
            return jsonify(
                {
                    "success": False,
                    "recognized": False,
                    "code": "scan_confirmation_required",
                    "message": (
                        "Please scan your face again, then choose today's "
                        "work mode."
                    ),
                }
            )

        employee = employee_by_code(employee_code)
        if not employee:
            clear_pending_scan_employee()
            return jsonify(
                {
                    "success": False,
                    "recognized": False,
                    "code": "employee_not_found",
                    "message": "Employee record could not be found.",
                }
            )

        result = process_scan_for_employee(employee, selected_work_mode)
        clear_pending_scan_employee()
        result["registered_faces"] = known_face_count()
        return jsonify(result)

    clear_pending_scan_employee()
    image_data = payload.get("image_data")
    if not image_data:
        return jsonify(
            {
                "success": False,
                "recognized": False,
                "code": "missing_image",
                "message": "No captured frame was received.",
            }
        )

    registered_faces = known_face_count()
    if registered_faces == 0:
        return jsonify(
            {
                "success": False,
                "recognized": False,
                "code": "no_registered_faces",
                "message": (
                    "No enrolled face profiles were found. "
                    "You should enroll first."
                ),
                "registered_faces": 0,
            }
        )

    try:
        _, face_encoding = extract_single_face_encoding(image_data)
    except (RuntimeError, ValueError) as exc:
        return jsonify(
            {
                "success": False,
                "recognized": False,
                "code": "scan_validation_failed",
                "message": str(exc),
                "registered_faces": registered_faces,
            }
        )

    employee = match_employee_from_encoding(face_encoding)
    if not employee:
        clear_pending_scan_employee()
        return jsonify(
            {
                "success": False,
                "recognized": False,
                "code": "face_not_found",
                "message": (
                    "This face is not enrolled yet. "
                    "You should enroll first."
                ),
                "registered_faces": registered_faces,
            }
        )

    current_record = today_record_for_employee(employee["employee_id"])
    if (
        current_record
        and current_record["check_in"]
        and current_record["check_out"]
    ):
        clear_pending_scan_employee()
        result = process_scan_for_employee(employee)
        result["registered_faces"] = registered_faces
        return jsonify(result)

    set_pending_scan_employee(employee["employee_id"])
    suggested_work_mode = normalize_work_mode(
        (
            current_record["work_mode"]
            if current_record and current_record["work_mode"]
            else employee["work_mode"]
        )
    )
    pending_action = (
        "check_out"
        if current_record and current_record["check_in"]
        else "check_in"
    )

    return jsonify(
        {
            "success": True,
            "recognized": True,
            "requires_work_mode": True,
            "message": (
                f"{employee['name']} recognized. "
                "Select today's work mode to continue."
            ),
            "employee": employee_view(employee),
            "pending_action": pending_action,
            "pending_action_label": (
                "Check Out Ready"
                if pending_action == "check_out"
                else "Check In Ready"
            ),
            "selected_work_mode": suggested_work_mode,
            "registered_faces": registered_faces,
        }
    )


@app.post("/api/enroll")
def api_enroll() -> Any:
    retry_after = consume_rate_limit(
        "enroll",
        max_requests=6,
        window_seconds=600,
    )
    if retry_after:
        return rate_limited_json_response(
            "Too many enrollment attempts. Please wait before trying again.",
            retry_after,
        )

    payload = request.get_json(silent=True) or {}
    if not payload:
        payload = request.form.to_dict()

    name, name_error = validate_employee_name(payload.get("name"))
    phone, phone_error = validate_phone_number(payload.get("phone"))
    department = normalize_department(payload.get("department"))
    work_mode = normalize_work_mode(payload.get("work_mode"))
    image_data = payload.get("image_data")

    if name_error or phone_error:
        return jsonify(
            {
                "success": False,
                "message": name_error or phone_error,
            }
        ), 400

    existing_employee = employee_by_enrollment_details(name, phone)
    if existing_employee:
        return jsonify(
            {
                "success": False,
                "message": (
                    f"You have already enrolled this employee as "
                    f"{existing_employee['name']} "
                    f"({existing_employee['employee_id']})."
                ),
            }
        ), 409

    if not image_data:
        return jsonify(
            {
                "success": False,
                "message": "Capture a face photo before enrolling.",
            }
        ), 400

    try:
        raw_bytes, face_encoding = extract_single_face_encoding(image_data)
    except (RuntimeError, ValueError) as exc:
        return jsonify({"success": False, "message": str(exc)}), 400

    existing_match = match_employee_from_encoding(face_encoding)
    if existing_match:
        return jsonify(
            {
                "success": False,
                "message": (
                    "This face is already enrolled as "
                    f"{existing_match['name']} "
                    f"({existing_match['employee_id']})."
                ),
            }
        ), 409

    employee_code = generate_employee_code()
    stored_photo_path = save_employee_photo(raw_bytes, employee_code)
    execute_db(
        """
        INSERT INTO employees (
            employee_id,
            name,
            email,
            phone,
            department,
            work_mode,
            password,
            role,
            face_encoding,
            photo_path,
            created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            employee_code,
            name,
            "",
            phone,
            department,
            work_mode,
            "face-enrolled",
            "employee",
            sqlite3.Binary(serialize_face_encoding(face_encoding)),
            stored_photo_path,
            iso_now(),
        ),
    )

    employee = employee_by_code(employee_code)
    assert employee is not None
    if employee["face_encoding"] is None:
        return jsonify(
            {
                "success": False,
                "message": (
                    "Enrollment did not save the face profile correctly. "
                    "Please try again."
                ),
            }
        ), 500

    return jsonify(
        {
            "success": True,
            "message": (
                f"{name} has been enrolled successfully. "
                "Face profile saved and ready for scanning."
            ),
            "employee": employee_view(employee),
            "registered_faces": known_face_count(),
        }
    )


@app.get("/api/today-stats")
def api_today_stats() -> Any:
    return jsonify(build_today_bundle())


@app.get("/api/export-csv")
@require_manager_access
def api_export_csv() -> Any:
    filters = history_filters_from_request(default_days=365)
    buffer = export_excel_rows(filters)
    filename = f"attendance-export-{today_ist().isoformat()}.xlsx"
    return send_file(
        buffer,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        as_attachment=True,
        download_name=filename,
    )


@app.errorhandler(404)
def not_found(_error: Exception) -> tuple[str, int]:
    return render_template("404.html"), 404


@app.errorhandler(500)
def server_error(_error: Exception) -> tuple[str, int]:
    return render_template("500.html"), 500


with app.app_context():
    ensure_schema()


def maybe_restart_with_project_venv() -> None:
    if not VENV_PYTHON.exists():
        return

    current_executable = Path(sys.executable).resolve()
    target_executable = VENV_PYTHON.resolve()
    if current_executable == target_executable:
        return

    if face_recognition is not None and cv2 is not None:
        return

    os.execv(str(target_executable), [str(target_executable), *sys.argv])


if __name__ == "__main__":
    maybe_restart_with_project_venv()
    app.run(
        host="0.0.0.0",
        port=int(os.getenv("PORT", "5000")),
        debug=os.getenv("FLASK_DEBUG", "0") == "1",
    )
